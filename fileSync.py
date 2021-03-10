#!/usr/bin/env python
# -*- coding:utf-8 -*-
import os
import time
import shutil
from openpyxl import Workbook
from openpyxl import load_workbook


def takeOrders(xlAddr):
    '''
    接收用户指令
    xlAddr：存储文件路径的 .xlsx 表格
    '''
    print('输入数字选择功能：1.同步文件；2.新增需要同步的文件：')
    orders = input()

    if(orders == "1"):
        # 1.同步文件
        print('====================')
        erroList = sync(xlAddr) # 开始同步
        
        # if(len(erroList)>0):
        #     print('部分同步失败，本地文件或网盘地址不存在：')
        #     for i in range(len(erroList)):
        #         print(erroList[i])

    if(orders == "2"):
        # 2.新增需要同步的文件
        print('====================')

        print('输入本地文件地址：')
        localFile = input()
        print('输入网盘地址：')
        networkDrive = input()
        
        # 二次确认
        # print('将 '+localFile+' 复制到 '+networkDrive+' ？输入 y 确认，n 重填')
        # ask = input()
        # if(ask == "y"):
        #     save(localFile, networkDrive, xlAddr)
        # else:
        #     print('----------')
        
        save(localFile, networkDrive, xlAddr)
        print('已保存: '+xlAddr)

    if(orders != "1" and orders != "2"):
        # 用户输入未知指令，重新询问指令
        takeOrders(xlAddr)


def save(localFile, networkDrive, xlAddr):
    '''
    处理用户输入：将保存到本地表格中
    localFile：原地址
    networkDrive：备份地址
    xlAddr：表格路径，上述地址将保存至此
    '''

    # print('====================')

    # 判断表格是否存在
    if(os.path.exists(xlAddr)):
        # 存在，打开本地数据表格
        wb = load_workbook(xlAddr)
    else:
        # 不存在，新建表格
        wb = Workbook()
    ws = wb.active  # 工作表

    dataTemp = []
    dataTemp.append(localFile) # 本地文件
    dataTemp.append(networkDrive) # 网盘地址
    dataTemp.append(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())) # 当前时间
    dataTemp.append(int(time.time())) # 时间戳

    ws.append(dataTemp)
    wb.save(xlAddr)


def sync(xlAddr):
    '''
    执行同步操作，将原文件拷贝到目标地址
    xlAddr：存储文件路径的 .xlsx 表格地址
    '''
    # print('====================')

    erroList = []

    # 判断表格是否存在
    if(os.path.exists(xlAddr)):
        # 存在，打开本地数据表格
        wb = load_workbook(xlAddr)
    else:
        # 不存在，新建表格
        wb = Workbook()
    ws = wb.active  # 工作表

    # 遍历表格，将本地文件拷贝到网盘地址
    for index in range(ws.max_row):
        index+=1

        if(index==1):
            # 跳过表头
            continue

        localFile = ws.cell(row=index, column=1).value
        networkDrive = ws.cell(row=index, column=2).value
        t = ws.cell(row=index, column=4).value
        
        # 过滤较旧的记录，只同步最近添加
        n = int(time.time())
        r = 60 # 同步最近 r 天的记录
        if(int(n)-int(t)>r*24*60*60):
            continue

        if(os.path.exists(localFile)==False):
            # 本地文件不存在
            erroList.append(localFile)
        if(os.path.exists(networkDrive)==False):
            # 网盘地址不存在
            erroList.append(networkDrive)
        if(os.path.exists(localFile)==False or os.path.exists(networkDrive)==False):
            print('！！！同步失败：'+str(index)+ ' '+localFile+' --> '+networkDrive)
            continue
        else:
            shutil.copyfile(localFile, networkDrive) # 拷贝到目标地址时，若存在同名文件则覆盖，不存在则直接复制
            print('同步成功：'+str(index)+ ' '+localFile+' --> '+networkDrive)

    return erroList

xlAddr = os.getcwd()+'/fileSync.xlsx' # .py 脚本同路径下的 .xlsx 文件
takeOrders(xlAddr)
input('操作完毕')